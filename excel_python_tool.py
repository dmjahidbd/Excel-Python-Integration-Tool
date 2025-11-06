"""Excel-Python Integration Tool
A comprehensive tool for seamless Excel and Python integration.
Allows users to automate data analysis, apply ML models, and perform advanced analytics.

Author: Your Name
GitHub: github.com/dmjahidbd/Excel-Python-Integration-Tool
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
import warnings
warnings.filterwarnings('ignore')

class ExcelPythonTool:
    """Main class for Excel-Python integration operations"""
    
    def __init__(self, file_path=None):
        """
        Initialize the Excel-Python Tool
        
        Args:
            file_path (str): Path to the Excel file
        """
        self.file_path = file_path
        self.df = None
        self.workbook = None
        
    def load_excel(self, file_path=None, sheet_name=0):
        """
        Load Excel file into pandas DataFrame
        
        Args:
            file_path (str): Path to Excel file
            sheet_name (str/int): Sheet name or index
            
        Returns:
            pd.DataFrame: Loaded data
        """
        if file_path:
            self.file_path = file_path
            
        if not self.file_path:
            raise ValueError("No file path provided")
            
        try:
            self.df = pd.read_excel(self.file_path, sheet_name=sheet_name)
            print(f"Successfully loaded {len(self.df)} rows and {len(self.df.columns)} columns")
            return self.df
        except Exception as e:
            print(f"Error loading Excel file: {str(e)}")
            return None
    
    def basic_analysis(self):
        """
        Perform basic statistical analysis on the data
        
        Returns:
            dict: Dictionary containing analysis results
        """
        if self.df is None:
            print("No data loaded. Please load an Excel file first.")
            return None
            
        analysis = {
            'shape': self.df.shape,
            'columns': list(self.df.columns),
            'dtypes': self.df.dtypes.to_dict(),
            'missing_values': self.df.isnull().sum().to_dict(),
            'numeric_summary': self.df.describe().to_dict()
        }
        
        return analysis
    
    def clean_data(self, drop_duplicates=True, fill_na_method='mean'):
        """
        Clean the data by removing duplicates and handling missing values
        
        Args:
            drop_duplicates (bool): Whether to drop duplicate rows
            fill_na_method (str): Method for filling NA values ('mean', 'median', 'mode', 'zero')
            
        Returns:
            pd.DataFrame: Cleaned data
        """
        if self.df is None:
            print("No data loaded.")
            return None
            
        # Drop duplicates
        if drop_duplicates:
            before_count = len(self.df)
            self.df = self.df.drop_duplicates()
            print(f"Removed {before_count - len(self.df)} duplicate rows")
        
        # Handle missing values
        if fill_na_method == 'mean':
            self.df = self.df.fillna(self.df.mean(numeric_only=True))
        elif fill_na_method == 'median':
            self.df = self.df.fillna(self.df.median(numeric_only=True))
        elif fill_na_method == 'zero':
            self.df = self.df.fillna(0)
        elif fill_na_method == 'mode':
            self.df = self.df.fillna(self.df.mode().iloc[0])
            
        return self.df
    
    def apply_formula(self, new_column_name, formula_func):
        """
        Apply a custom formula to create a new column
        
        Args:
            new_column_name (str): Name of the new column
            formula_func (callable): Function to apply to each row
            
        Returns:
            pd.DataFrame: Updated dataframe
        """
        if self.df is None:
            print("No data loaded.")
            return None
            
        try:
            self.df[new_column_name] = self.df.apply(formula_func, axis=1)
            print(f"Successfully created column: {new_column_name}")
            return self.df
        except Exception as e:
            print(f"Error applying formula: {str(e)}")
            return None
    
    def filter_data(self, condition):
        """
        Filter data based on a condition
        
        Args:
            condition: Boolean condition for filtering
            
        Returns:
            pd.DataFrame: Filtered data
        """
        if self.df is None:
            print("No data loaded.")
            return None
            
        filtered_df = self.df[condition]
        print(f"Filtered data: {len(filtered_df)} rows")
        return filtered_df
    
    def aggregate_data(self, group_by_columns, agg_dict):
        """
        Aggregate data by grouping
        
        Args:
            group_by_columns (list): Columns to group by
            agg_dict (dict): Aggregation functions
            
        Returns:
            pd.DataFrame: Aggregated data
        """
        if self.df is None:
            print("No data loaded.")
            return None
            
        try:
            aggregated = self.df.groupby(group_by_columns).agg(agg_dict).reset_index()
            return aggregated
        except Exception as e:
            print(f"Error aggregating data: {str(e)}")
            return None
    
    def export_to_excel(self, output_path, data=None, sheet_name='Sheet1', styled=False):
        """
        Export data to Excel with optional styling
        
        Args:
            output_path (str): Path for output Excel file
            data (pd.DataFrame): Data to export (uses self.df if None)
            sheet_name (str): Name of the sheet
            styled (bool): Whether to apply styling
            
        Returns:
            bool: Success status
        """
        if data is None:
            data = self.df
            
        if data is None:
            print("No data to export.")
            return False
            
        try:
            if styled:
                # Export with styling
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]
                    
                    # Style header row
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFF')
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            else:
                # Simple export
                data.to_excel(output_path, sheet_name=sheet_name, index=False)
            
            print(f"Successfully exported to {output_path}")
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {str(e)}")
            return False
    
    def pivot_analysis(self, index, columns, values, aggfunc='sum'):
        """
        Create pivot table analysis
        
        Args:
            index: Column to use as index
            columns: Column to use as columns
            values: Column to aggregate
            aggfunc: Aggregation function
            
        Returns:
            pd.DataFrame: Pivot table
        """
        if self.df is None:
            print("No data loaded.")
            return None
            
        try:
            pivot = pd.pivot_table(self.df, index=index, columns=columns, 
                                 values=values, aggfunc=aggfunc)
            return pivot
        except Exception as e:
            print(f"Error creating pivot table: {str(e)}")
            return None


def demo_usage():
    """
    Demonstration of tool usage with sample data
    """
    print("=" * 60)
    print("Excel-Python Integration Tool - Demo")
    print("=" * 60)
    
    # Create sample data
    sample_data = {
        'Product': ['A', 'B', 'C', 'A', 'B', 'C', 'A', 'B'],
        'Region': ['East', 'East', 'West', 'West', 'East', 'East', 'West', 'West'],
        'Sales': [100, 150, 200, 120, 180, 210, 110, 160],
        'Quantity': [10, 15, 20, 12, 18, 21, 11, 16]
    }
    
    df = pd.DataFrame(sample_data)
    
    # Save sample data
    sample_file = 'sample_data.xlsx'
    df.to_excel(sample_file, index=False)
    print(f"\n✓ Created sample data: {sample_file}")
    
    # Initialize tool
    tool = ExcelPythonTool(sample_file)
    
    # Load data
    print("\n1. Loading Excel file...")
    tool.load_excel()
    
    # Basic analysis
    print("\n2. Performing basic analysis...")
    analysis = tool.basic_analysis()
    print(f"   Shape: {analysis['shape']}")
    print(f"   Columns: {analysis['columns']}")
    
    # Apply custom formula
    print("\n3. Applying custom formula (Revenue = Sales * Quantity)...")
    tool.apply_formula('Revenue', lambda row: row['Sales'] * row['Quantity'])
    
    # Aggregate data
    print("\n4. Aggregating data by Product...")
    aggregated = tool.aggregate_data(['Product'], {
        'Sales': 'sum',
        'Quantity': 'sum',
        'Revenue': 'sum'
    })
    print(aggregated)
    
    # Export results
    print("\n5. Exporting results with styling...")
    tool.export_to_excel('output_styled.xlsx', styled=True)
    tool.export_to_excel('output_aggregated.xlsx', data=aggregated, styled=True)
    
    print("\n" + "=" * 60)
    print("Demo completed! Check the generated Excel files.")
    print("=" * 60)


if __name__ == "__main__":
    # Run demo
    demo_usage()
